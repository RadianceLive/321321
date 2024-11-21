import pandas as pd
import requests
import random
from bs4 import BeautifulSoup
# import json
from openpyxl import *
from fake_headers import Headers
# import cloudscraper
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import undetected_chromedriver as us
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
import time
import math
import chromedriver_autoinstaller

def parse():
    wb = Workbook()
    chromedriver_autoinstaller.install()
    driver = webdriver.Chrome()
    
    main_url = ['https://eats.yandex.com/ru-am/Yerevan?quickfilter=sushi&shippingType=delivery']
    
    
    driver.get(f'https://eats.yandex.com/ru-am/Yerevan?quickfilter=sushi&shippingType=delivery')
    driver.get(f'https://eats.yandex.com/ru-am/Yerevan?quickfilter=sushi&shippingType=delivery')
    element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="address-button-root"]'))
)
    # input('OK: ')
    time.sleep(2)
    # Кликаем на объект
    element.click()
    element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="address-button-add"]'))
)
    
    time.sleep(2)
    # Кликаем на объект
    element.click()
    element = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="address-input-reset"]'))
)
    
    time.sleep(2)
    # Кликаем на объект
    element.click()
    input_text = "улица Геворка Джаукяна, 53"
    try:
        # Ожидание появления элемента input
        input_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".AppAddressInput_addressInput"))
        )
        # Ввод текста
        input_element.send_keys(input_text)
        driver.find_element(By.CSS_SELECTOR, ".AppAddressInput_addressInput").send_keys(Keys.ENTER)

        # Находим кнопку ОК и кликаем по ней
        button_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-testid="desktop-location-modal-confirm-button"]'))
        )
        # time.sleep(1)
        button_element.click()
        

    except:
        # Закрытие браузера после выполнения
        input('Ok: ')

    time.sleep(2)
    scroll_height = driver.execute_script("return document.body.scrollHeight")

    # Шаг скроллинга
    scroll_step = 1500

    # Цикл скроллинга страницы
    for i in range(0, scroll_height, scroll_step):
        # driver.execute_script(f"window.scrollTo(0, {i})")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1)  # Задержка для плавного скроллинга (меньше = быстрее)

    hrefs = driver.find_elements(By.CSS_SELECTOR,'[data-testid="desktop-place-item"]')
    names = driver.find_elements(By.CSS_SELECTOR,'[data-testid="desktop-place-item-name"]')
    # print(names)
    arr = []
    arr_names = []
    for item in hrefs:
        arr.append(item.get_attribute("href"))
    for item in names:
        arr_names.append(item.text)
    
    # print(arr_names)

    # перебор товаров в рестике
    for idx, href in enumerate(arr):
        c = 2    
        driver.get(href)
        
        name = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[class="RestaurantHeader_name UiKitText_root UiKitText_HeaderLoose UiKitText_Bold UiKitText_Text"]'))
        )
            
        
        time.sleep(10)
        categories_body =driver.page_source
        cates = BeautifulSoup(categories_body, 'html.parser').find_all(class_='RestaurantCategories_item')[2:]

        arr_cat = []
        for item in cates:
            arr_cat.append(item.text)
        # input('ok:')
        
        scroll_height = driver.execute_script("return document.body.scrollHeight")

        # Шаг скроллинга
        scroll_step = 800
        arr_pos = []
        # Цикл скроллинга страницы
        for i in range(0, scroll_height, scroll_step):
            driver.execute_script(f"window.scrollTo(0, {i})")
            # driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1) 
            categories_body =driver.page_source

            pos_dis = BeautifulSoup(categories_body, 'html.parser').select('[data-testid="product-card-root"]')
            pos_name = BeautifulSoup(categories_body, 'html.parser').find_all(class_='UiKitDesktopProductCard_name UiKitText_root UiKitText_Body2Loose UiKitText_Regular UiKitText_Text')
            pos_cost = BeautifulSoup(categories_body, 'html.parser').find_all(class_='UiKitDesktopProductCard_priceWrapper')
            pos_pict = BeautifulSoup(categories_body, 'html.parser').select('[data-testid="smart-image-img"]')[1:]
            # print(len(pos_dis))

            for i in range(len(pos_name)):
                if f'{pos_name[i].text} | {pos_cost[i].text} | {pos_dis[i].get("description")} | {pos_pict[i].get("src")}' in arr_pos:
                    pass
                else:
                    arr_pos.append(f'{pos_name[i].text} | {pos_cost[i].text} | {pos_dis[i].get("description")} | {pos_pict[i].get("src")}')
                    # print(f'{pos_name[i].text} | {pos_cost[i].text} | {pos_dis[i].get("description")} | {pos_pict[i].get("src")}')
            

        for url in main_url:
            file = f'restaurant_.xlsx'
            # try:
            #     xl = pd.ExcelFile(file)
            # except:
            # wb.save(file)
            # print(len(arr_names))
            wb.create_sheet(f'{arr_names[idx+1] if len(arr_names[idx+1]) <= 30 else arr_names[idx+1][:7]}')
            wb.save(file)
            xl = pd.ExcelFile(file)
            # xl = pd.ExcelFile(file)
            sheet = wb[xl.sheet_names[idx+1]]
            sheet[f'A1'].value = "Название товара"
            sheet[f'B1'].value = "Цена"
            sheet[f'C1'].value = "Описание"
            sheet[f'D1'].value = "Картинка"
            sheet[f'E1'].value = "Название ресторана"
            sheet[f'F1'].value = "Категории"
            # sheet[f'G1'].value = "Адрес"
            # sheet[f'H1'].value = "График работы"
            # sheet[f'I1'].value = "Телефон"
            # sheet[f'J1'].value = "Сайт"
            

            sheet[f'A1'].style = "Headline 2"
            sheet[f'B1'].style = "Headline 2"
            sheet[f'C1'].style = "Headline 2"
            sheet[f'D1'].style = "Headline 2"
            sheet[f'E1'].style = "Headline 2"
            sheet[f'F1'].style = "Headline 2"
            # sheet[f'G1'].style = "Headline 2"
            # sheet[f'H1'].style = "Headline 2"
            # sheet[f'I1'].style = "Headline 2"
            # sheet[f'J1'].style = "Headline 2"
            # wb.save(file)
        # try:
        #     wb = load_workbook(f'restaurant_{url.split("=")[2]}.xlsx')
        #     sheet = wb[xl.sheet_names[0]]
        # except:
        #     pass
        # count = len(sheet['A']) + 1
        # all_rows = len(sheet['A'])
            for items in arr_pos:
                name_pos = items.split(' | ')[0]
                cost_pos = items.split(' | ')[1]
                dis_pos = items.split(' | ')[2]
                pict_pos = items.split(' | ')[3]
                name_rest = name.text
                cats = ', '.join(cat for cat in arr_cat)
                sheet[f'A{c}'].value = name_pos
                sheet[f'B{c}'].value = cost_pos
                sheet[f'C{c}'].value = dis_pos
                sheet[f'D{c}'].value = pict_pos
                sheet[f'E{c}'].value = name_rest
                sheet[f'F{c}'].value = cats
                c = c + 1
                wb.save(file)
                print("\x1bc\x1b[H",end="")
                print(f'{c-2} / {len(arr_pos)} / {name_rest} / {len(hrefs)}')
                # print(f'{pos_name[i].text} | {pos_cost[i].text} | {pos_dis[i].get("description")} | {pos_pict[i].get("src")}')








    # time.sleep(60)    

        # scroll_height = driver.execute_script("return document.body.scrollHeight")
        print(name.text)
        print(arr_cat)
        # if idx == 1:
            #     break
        # for i in len(hrefs):


parse()



